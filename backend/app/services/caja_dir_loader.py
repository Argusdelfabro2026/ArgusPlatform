# app/services/caja_dir_loader.py
# Loads Caja Digital / Caja Dirección Excel and filters by canal == 1 (Transfer).
#
# Expected column format (from client's Caja Digital file):
#   DIA | Fecha | NRO TIPO | TIPO | IMPORTE | DESCRIPCIÓN | usd | C | Cliente
#   | C2 (canal numeric) | Canal (text) | importe2 (signed amount) | SEMANA
#
# Key mapping:
#   C2        → canal field   (filter: C2 == 1 = Transfer)
#   TIPO      → category      (has leading space, e.g. " VENTAS" — stripped on read)
#   importe2  → signed amount (negative = expense, positive = income)
#   Fecha     → date

import logging
from datetime import date
from typing import Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger("argus.caja_dir")

# Standard column name aliases (case-insensitive exact match after strip)
_FECHA_NAMES     = {"fecha", "date", "fecha_movimiento", "fecha_mov", "fecha mov"}
_CATEGORIA_NAMES = {"tipo", "categoria", "categoría", "concepto", "rubro"}
_DESC_NAMES      = {"descripcion", "descripción", "description", "detalle", "glosa"}
_IMPORTE_NAMES   = {"importe", "monto", "amount", "total", "haber", "debe"}
_CANAL_NAMES     = {"canal", "channel", "tipo_canal", "tipo canal"}

# Priority overrides: these column names always win over the generic aliases above.
# Key = exact lowercase column header → logical field name to assign.
_PRIORITY_COLS: Dict[str, str] = {
    "importe2": "importe",  # signed amount — preferred over plain "importe"
    "c2":       "canal",    # numeric canal (1=Transfer) — preferred over text "Canal"
}


def _detect_columns(header_row: tuple) -> Dict[str, int]:
    """Map logical field names to column indices from a header tuple."""
    mapping: Dict[str, int] = {}
    # Tracks fields already claimed by a priority column so duplicate tables
    # (e.g. the same header appearing twice in one sheet) don't overwrite the
    # first — correct — occurrence.
    priority_mapped: set = set()

    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()

        # Priority columns override generic aliases but only on first occurrence.
        if name in _PRIORITY_COLS:
            field = _PRIORITY_COLS[name]
            if field not in priority_mapped:
                mapping[field] = idx
                priority_mapped.add(field)
            continue

        # Standard aliases — first match wins (only if not already set by priority)
        if name in _FECHA_NAMES and "fecha" not in mapping:
            mapping["fecha"] = idx
        if name in _CATEGORIA_NAMES and "categoria" not in mapping:
            mapping["categoria"] = idx
        if name in _DESC_NAMES and "descripcion" not in mapping:
            mapping["descripcion"] = idx
        if name in _IMPORTE_NAMES and "importe" not in mapping and "importe" not in priority_mapped:
            mapping["importe"] = idx
        if name in _CANAL_NAMES and "canal" not in mapping and "canal" not in priority_mapped:
            mapping["canal"] = idx

    return mapping


def _load_sheet_rows(
    sheet_name: str,
    all_rows: list,
    warnings: List[str],
) -> Tuple[List[dict], int, int, Dict[str, int]]:
    """
    Process one sheet's rows: detect columns, filter canal=1, return parsed rows.

    Returns (rows, total, skipped, canal_counts).
    """
    # Find first non-empty row as header
    header_row: Optional[tuple] = None
    data_start = 0
    for i, row in enumerate(all_rows):
        if any(cell is not None for cell in row):
            header_row = row
            data_start = i + 1
            break

    if header_row is None:
        warnings.append(f"Hoja '{sheet_name}': sin encabezado — omitida")
        return [], 0, 0, {}

    col_map = _detect_columns(header_row)
    logger.info(f"Caja Digital [{sheet_name}] — columnas detectadas: {col_map}")

    if "importe" not in col_map:
        warnings.append(f"Hoja '{sheet_name}': sin columna importe/importe2 — omitida")
        return [], 0, 0, {}
    if "categoria" not in col_map:
        warnings.append(f"Hoja '{sheet_name}': sin columna categoría/TIPO — se usará vacío")
    if "canal" not in col_map:
        warnings.append(f"Hoja '{sheet_name}': sin columna C2/Canal — se incluirán todos los registros")

    rows: List[dict] = []
    total = 0
    skipped = 0
    canal_counts: Dict[str, int] = {}

    for row in all_rows[data_start:]:
        if all(cell is None for cell in row):
            continue
        total += 1

        if "canal" in col_map:
            canal_val = row[col_map["canal"]]
            canal_key = str(canal_val) if canal_val is not None else "None"
            canal_counts[canal_key] = canal_counts.get(canal_key, 0) + 1
            try:
                if int(canal_val) != 1:
                    skipped += 1
                    continue
            except (TypeError, ValueError):
                skipped += 1
                continue

        fecha_raw = row[col_map["fecha"]] if "fecha" in col_map else None
        fecha: Optional[date] = None
        if isinstance(fecha_raw, date):
            fecha = fecha_raw
        elif hasattr(fecha_raw, "date"):
            fecha = fecha_raw.date()

        cat_idx = col_map.get("categoria")
        categoria = (
            str(row[cat_idx]).strip()
            if (cat_idx is not None and row[cat_idx] is not None)
            else ""
        )

        importe = 0.0
        try:
            raw = row[col_map["importe"]]
            if raw is not None:
                importe = float(raw)
        except (TypeError, ValueError):
            pass

        desc_idx = col_map.get("descripcion")
        descripcion = (
            str(row[desc_idx]).strip()
            if (desc_idx is not None and row[desc_idx] is not None)
            else ""
        )

        rows.append({"fecha": fecha, "categoria": categoria, "importe": importe, "descripcion": descripcion})

    return rows, total, skipped, canal_counts


def load_caja_direccion(path: str) -> Tuple[List[dict], List[str]]:
    """
    Load Caja Digital / Caja Dirección Excel and filter by canal == 1 (Transfer).

    Reads ALL sheets in the workbook (one per month) and consolidates the results.

    Returns:
        rows     — list of dicts: {fecha, categoria (stripped), importe (signed)}
        warnings — non-fatal issues detected during load
    """
    warnings: List[str] = []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"No se pudo abrir el archivo de Caja: {exc}"]

    sheet_names = wb.sheetnames
    if not sheet_names:
        return [], ["El archivo de Caja no tiene hojas"]

    logger.info(f"Caja Digital — {len(sheet_names)} hojas encontradas: {sheet_names}")

    all_rows: List[dict] = []
    total_all = 0
    skipped_all = 0
    canal_counts_all: Dict[str, int] = {}
    sheets_loaded = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows or all(all(c is None for c in r) for r in sheet_rows):
            logger.info(f"Caja Digital [{sheet_name}] — vacía, omitida")
            continue

        rows, total, skipped, canal_counts = _load_sheet_rows(sheet_name, sheet_rows, warnings)

        all_rows.extend(rows)
        total_all  += total
        skipped_all += skipped
        for k, v in canal_counts.items():
            canal_counts_all[k] = canal_counts_all.get(k, 0) + v

        logger.info(
            f"Caja Digital [{sheet_name}] — total: {total}, "
            f"filtrados (canal≠1): {skipped}, Transfer: {len(rows)}"
        )
        sheets_loaded += 1

    if sheets_loaded == 0:
        return [], ["No se encontraron hojas con datos válidos en el archivo de Caja"]

    if canal_counts_all:
        dist = ", ".join(f"canal={k}: {v}" for k, v in sorted(canal_counts_all.items()))
        logger.info(
            f"Caja Digital — TOTAL {sheets_loaded} hojas: {total_all} filas, "
            f"filtradas: {skipped_all}, Transfer: {len(all_rows)} | [{dist}]"
        )
        warnings.append(f"__CANAL_DIST__ [{dist}] → {len(all_rows)} Transfer rows kept ({sheets_loaded} sheets)")
    else:
        logger.info(f"Caja Digital — total procesados: {len(all_rows)} ({sheets_loaded} hojas)")

    return all_rows, warnings
